import os
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

VIDEO_EXTENSIONS = ['.mp4', '.mkv', '.mov', '.avi', '.webm']

def get_duration(file_path):
    try:
        result = subprocess.run(
            ['ffprobe', '-v', 'error', '-show_entries',
             'format=duration', '-of',
             'default=noprint_wrappers=1:nokey=1', file_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT
        )
        return float(result.stdout)
    except:
        return 0.0

def format_duration(seconds):
    minutes = seconds / 60
    return round(minutes, 2)

def minutes_to_hours_minutes(minutes):
    hours = int(minutes // 60)
    mins = int(minutes % 60)
    return f"{hours}h {mins}m"

def scan_directory(root_dir):
    data = []  # list of tuples (section, subsection, duration_in_min)
    for root, dirs, files in os.walk(root_dir):
        section = os.path.basename(root)
        if section == os.path.basename(root_dir):
            continue

        for file in files:
            if os.path.splitext(file)[1].lower() in VIDEO_EXTENSIONS:
                file_path = os.path.join(root, file)
                duration = get_duration(file_path)
                data.append((section, file, format_duration(duration)))

    return data

def create_excel(data, filename="course_progress.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Progress"

    headers = ["Section", "Subsection", "Duration (min)", "Watched (0/1)"]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data rows
    for i, (section, subsection, duration) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=section)
        ws.cell(row=i, column=2, value=subsection)
        ws.cell(row=i, column=3, value=duration)
        ws.cell(row=i, column=4, value=0)  # Default watched = 0

    max_row = ws.max_row

    # Add data validation for watched column (only 0 or 1)
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    ws.add_data_validation(dv)
    for row in range(2, max_row + 1):
        dv.add(ws.cell(row=row, column=4))

    # Group rows by section
    section_rows = {}
    section_durations = {}
    for row in range(2, max_row+1):
        section = ws.cell(row=row, column=1).value
        dur = ws.cell(row=row, column=3).value
        section_rows.setdefault(section, []).append(row)
        section_durations[section] = section_durations.get(section, 0) + dur

    # Calculate total duration of the course
    total_duration = sum(section_durations.values())

    # Summary table headers below data
    summary_start = max_row + 3
    ws.cell(row=summary_start, column=1, value="Section").font = Font(bold=True)
    ws.cell(row=summary_start, column=2, value="Progress (%)").font = Font(bold=True)
    ws.cell(row=summary_start, column=3, value="Time Watched (min)").font = Font(bold=True)
    ws.cell(row=summary_start, column=4, value="Time Remaining (min)").font = Font(bold=True)
    ws.cell(row=summary_start, column=5, value="Total Time (min)").font = Font(bold=True)
    ws.cell(row=summary_start, column=6, value="Total Time (h:m)").font = Font(bold=True)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Write each section's progress and time info
    current_row = summary_start + 1
    for section, rows in section_rows.items():
        # Progress formula (average of watched column)
        watched_range = f"D{rows[0]}:D{rows[-1]}"
        formula_progress = f"=AVERAGE({watched_range})"
        ws.cell(row=current_row, column=1, value=section)
        ws.cell(row=current_row, column=2, value=formula_progress)
        ws.cell(row=current_row, column=2).number_format = '0.00%'

        # Time watched: SUMPRODUCT(Duration * Watched)
        first_row = rows[0]
        last_row = rows[-1]
        formula_time_watched = f"=SUMPRODUCT(C{first_row}:C{last_row},D{first_row}:D{last_row})"
        ws.cell(row=current_row, column=3, value=formula_time_watched)

        # Time remaining = total duration - watched
        total_section_time = section_durations[section]
        ws.cell(row=current_row, column=5, value=round(total_section_time, 2))

        # Formula for remaining time: total - watched
        formula_time_remaining = f"=E{current_row}-C{current_row}"
        ws.cell(row=current_row, column=4, value=formula_time_remaining)

        # Total time in h:m (calculated in Python)
        ws.cell(row=current_row, column=6, value=minutes_to_hours_minutes(total_section_time))

        # Apply border
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = thin_border

        current_row += 1

    # Total row for whole course
    ws.cell(row=current_row, column=1, value="Total")
    ws.cell(row=current_row, column=2, value=f"=AVERAGE(D2:D{max_row})")
    ws.cell(row=current_row, column=2).number_format = '0.00%'

    ws.cell(row=current_row, column=3, value=f"=SUMPRODUCT(C2:C{max_row},D2:D{max_row})")
    ws.cell(row=current_row, column=5, value=round(total_duration, 2))

    ws.cell(row=current_row, column=4, value=f"=E{current_row}-C{current_row}")
    ws.cell(row=current_row, column=6, value=minutes_to_hours_minutes(total_duration))

    for col in range(1, 7):
        ws.cell(row=current_row, column=col).border = thin_border
        ws.cell(row=current_row, column=col).font = Font(bold=True)

    # Conditional formatting for data rows:
    # If average watched for section = 1 -> green fill, else red fill for those rows
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for section, rows in section_rows.items():
        avg_formula = f"=AVERAGE($D${rows[0]}:$D${rows[-1]})=1"
        avg_formula_red = f"=AVERAGE($D${rows[0]}:$D${rows[-1]})<1"

        for row in rows:
            for col in range(1, 5):
                ws.conditional_formatting.add(
                    f"{get_column_letter(col)}{row}",
                    FormulaRule(formula=[avg_formula], fill=green_fill)
                )
                ws.conditional_formatting.add(
                    f"{get_column_letter(col)}{row}",
                    FormulaRule(formula=[avg_formula_red], fill=red_fill)
                )

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15

    # Center align watched column
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

    print(f"Excel file '{filename}' created successfully.")
    wb.save(filename)

if __name__ == "__main__":
    folder_path = os.getcwd()
    data = scan_directory(folder_path)
    create_excel(data)
